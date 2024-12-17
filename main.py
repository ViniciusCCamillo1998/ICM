import pandas as pd
import numpy as np
import os
import shutil
import openpyxl as op
from copy import copy
from tkinter import filedialog, messagebox

class CalcICM:
    def __init__(self, file_path, step):
        self.file_path = file_path
        self.step = int(round(step*1000, 3)) # Convert to m
        self.list_df_main = []

    def SeparaDF(self, df):
        # Criar uma lista de dataframes separados
        dfs_separados = []

        # Iterar sobre as linhas do dataframe original
        old_index = 0
        for index, row in df.iterrows():
            # Verificar se a linha está em branco
            if row.isnull().all():
                dfs_separados.append(df.loc[old_index:index-1])
                old_index = index + 1

        # Adicionar o último dataframe atual à lista de dataframes separados
        dfs_separados.append(df.loc[old_index:index])
        #print(dfs_separados)

        return dfs_separados

    def Sections(self, df):
        # Original
        start_raw = int(round(df['Início'].tolist()[0]*1000, 3))
        end_raw = int(round(df['Fim'].tolist()[-1]*1000, 3))
        if start_raw - end_raw < 0:
            # Crescent
            start = int((round(start_raw/self.step, 3)) + 1)*self.step
            segments = list(range(int(start), int(end_raw), self.step))
        else:
            # Descending
            start = int(round(start_raw/self.step, 3))*self.step
            segments = list(range(int(start), int(end_raw), -self.step))

        if len(segments) == 0:
            segments.insert(0, start_raw)
        if segments[0] != start_raw:
            segments.insert(0, start_raw)
        if segments[-1] != end_raw:
            segments.append(end_raw)

        return segments

    def GetHeader(self, df):
        df_original = pd.read_excel(self.file_path, usecols = "A:B")
        for coluna in df_original.columns:
            if coluna != "Rodovia":
                road = str(coluna)
        segment_name = os.path.basename(self.file_path)
        one_way = "X" if segment_name.split("_")[2] == "S" else ""
        two_way = "X" if segment_name.split("_")[2] == "D" else ""

        line_step = round(abs(round(df['Fim'].tolist()[0], 3) - round(df['Fim'].tolist()[1], 3))*1000, 3)

        nome_trecho = segment_name.split("_")[1]
        tipo_pista = str(df_original.iat[0, 1])
        faixa = str(df_original.iat[1, 1])
        data_bruta = df_original.iat[4, 1]
        try:
            data = ("{}/{}/{}".format(data_bruta.day, data_bruta.month, data_bruta.year))
        except:
            data = data_bruta
        km_ini = float(df_original.iat[2, 1])
        km_fim = float(df_original.iat[3, 1])
        sentido_pista = 'Crescente' if km_ini - km_fim < 0 else 'Decrescente'

        return {'Road': road, "STH":nome_trecho, "Track":tipo_pista, "Way":sentido_pista, "Lane":faixa, "Data":data,
                "km ini":km_ini, "km fim":km_fim, 'one way': one_way, 'two way': two_way, 'Line Step': line_step}
    
    def Processing(self, df_base=pd.DataFrame()):
        # Creating final df
        df_final = pd.DataFrame(columns=["Rodovia", "Início", "Fim", "Extensão", "P Alto", "P Médio", "P Baixo", "R Alto", "R Médio", "R Baixo", "Tr Alto", "Tr Médio",
                                        "Tr Baixo", "Roçada.B", "Roçada.M", "Roçada.R", "Drenagem.B", "Drenagem.M", "Drenagem.R", "Sinalização.B", "Sinalização.M",
                                        "Sinalização.R", "Data", "Latitude", "Longitude", "Observação", "ICC", "ICP", "ICM"])
        
        for df_list in self.list_df_main:
            df_list.reset_index(drop=True, inplace=True)
            df = df_list.copy(deep=True)

            segments = self.Sections(df)
            file_header = self.GetHeader(df)

            if not df_base.empty:
                # Crescent
                if segments[0] - segments[-1] < 0:
                    df_temp_base = df_base[(df_base['Início'] >= round(segments[0]/1000, 3)) & (df_base['Fim'] <= round(segments[-1]/1000, 3))]
                # Descending
                else:
                    df_temp_base = df_base[(df_base['Início'] <= round(segments[0]/1000, 3)) & (df_base['Fim'] >= round(segments[-1]/1000, 3))]

                df_temp_base.reset_index(drop=True, inplace=True)
                for col in ["Roçada.B", "Roçada.M", "Roçada.R", "Drenagem.B", "Drenagem.M", "Drenagem.R", "Sinalização.B", "Sinalização.M", "Sinalização.R"]:
                    df[col] = df_temp_base[col]

            for segment in range(len(segments) - 1):
                # Cleaning temporary dictionary
                dict_segment = {column: [None] for column in ["Rodovia", "Início", "Fim", "Extensão", "P Alto", "P Médio", "P Baixo", "R Alto", "R Médio", "R Baixo",
                                                               "Tr Alto", "Tr Médio", "Tr Baixo", "Roçada.B", "Roçada.M", "Roçada.R", "Drenagem.B", "Drenagem.M",
                                                               "Drenagem.R", "Sinalização.B", "Sinalização.M", "Sinalização.R", "Data", "Latitude", "Longitude", 
                                                               "Observação", "ICC", "ICP", "ICM"]}
                # Crescent
                if segments[0] - segments[-1] < 0:
                    df_temp = df[(df['Início'] >= round(segments[segment]/1000, 3)) & (df['Fim'] <= round(segments[segment + 1]/1000, 3))]
                # Descending
                else:
                    df_temp = df[(df['Início'] <= round(segments[segment]/1000, 3)) & (df['Fim'] >= round(segments[segment + 1]/1000, 3))]
                
                # Count
                P_count = df_temp['P'].count()
                R_count = df_temp['R'].count()
                Tr_area = file_header['Line Step'] * (((df_temp['Tr.BE'].count() + df_temp['Tr.BD'].count()))*0.4 + ((df_temp['Tr.ATRE'].count() + df_temp['Tr.ATRD'].count()))*0.8 + df_temp['Tr.F'].count()*3.6)

                # Filling in dict
                dict_segment['Rodovia'] = [file_header['Road']]
                dict_segment['Início'] = [round(segments[segment]/1000, 3)]
                dict_segment['Fim'] = [round(segments[segment + 1]/1000, 3)]
                dict_segment['Extensão'] = [round(abs(segments[segment] - segments[segment + 1])/1000, 3)]

                dict_segment['Data'] = df_temp['Data'].tolist()[0]
                dict_segment['Latitude'] = df_temp['Latitude'].tolist()[0]
                dict_segment['Longitude'] = df_temp['Longitude'].tolist()[0]

                obs = ""
                for observacao in df_temp['Observação']:
                    if obs and not pd.isna(observacao):
                        obs=obs+" | "+str(observacao)
                    elif not pd.isna(observacao):
                        obs=str(observacao)
                dict_segment['Observação'] = [obs]

                # Filling in Pothole
                if P_count <= 2:
                    dict_segment['P Baixo'] = ["X"]
                    fc_p = 0.25
                elif P_count > 2 and P_count <= 5:
                    dict_segment['P Médio'] = ["X"]
                    fc_p = 0.5
                else:
                    dict_segment['P Alto'] = ["X"]
                    fc_p = 1
                
                # Filling in Patch
                if R_count <= 2:
                    dict_segment['R Baixo'] = ["X"]
                    fc_r = 0.25
                elif R_count > 2 and R_count <= 5:
                    dict_segment['R Médio'] = ["X"]
                    fc_r = 0.5
                else:
                    dict_segment['R Alto'] = ["X"]
                    fc_r = 1
                
                # Filling in crack
                Tr_percent = (Tr_area/(dict_segment['Extensão'][0]*3.6*1000))*100
                if Tr_percent <= 10:
                    dict_segment['Tr Baixo'] = ["X"]
                    fc_tr = 0.25
                elif Tr_percent > 10 and Tr_percent <= 50:
                    dict_segment['Tr Médio'] = ["X"]
                    fc_tr = 0.5
                else:
                    dict_segment['Tr Alto'] = ["X"]
                    fc_tr = 1

                dict_segment['ICP'] = [(50*fc_p) + (30*fc_r) + (20*fc_tr)]

                # Filling in mowing
                mowing_b = df_temp['Roçada.B'].count()*0.25
                mowing_m = df_temp['Roçada.M'].count()*0.5
                mowing_r = df_temp['Roçada.R'].count()*1
                if mowing_b > mowing_m and mowing_b > mowing_r:
                    dict_segment['Roçada.B'] = ["X"]
                    fc_mowing = 0.25
                elif mowing_m > mowing_b and mowing_m > mowing_r:
                    dict_segment['Roçada.M'] = ["X"]
                    fc_mowing = 0.5
                elif mowing_r > mowing_b and mowing_r > mowing_m:
                    dict_segment['Roçada.R'] = ["X"]
                    fc_mowing = 1
                elif mowing_b == 0 and mowing_m == 0 and mowing_r == 0:
                    dict_segment['Roçada.B'] = ["X"]
                    fc_mowing = 0.25

                # Filling in drainage
                drainage_b = df_temp['Drenagem.B'].count()*0.25
                drainage_m = df_temp['Drenagem.M'].count()*0.5
                drainage_r = df_temp['Drenagem.R'].count()*1
                if drainage_b > drainage_m and drainage_b > drainage_r:
                    dict_segment['Drenagem.B'] = ["X"]
                    fc_drainage = 0.25
                elif drainage_m > drainage_b and drainage_m > drainage_r:
                    dict_segment['Drenagem.M'] = ["X"]
                    fc_drainage = 0.5
                elif drainage_r > drainage_b and drainage_r > drainage_m:
                    dict_segment['Drenagem.R'] = ["X"]
                    fc_drainage = 1
                elif drainage_b == 0 and drainage_m == 0 and drainage_r == 0:
                    dict_segment['Drenagem.B'] = ["X"]
                    fc_drainage = 0.25
                
                # Filling in signaling
                signaling_b = df_temp['Sinalização.B'].count()*0.25
                signaling_m = df_temp['Sinalização.M'].count()*0.5
                signaling_r = df_temp['Sinalização.R'].count()*1
                if signaling_b > signaling_m and signaling_b > signaling_r:
                    dict_segment['Sinalização.B'] = ["X"]
                    fc_signaling = 0.25
                elif signaling_m > signaling_b and signaling_m > signaling_r:
                    dict_segment['Sinalização.M'] = ["X"]
                    fc_signaling = 0.5
                elif signaling_r > signaling_b and signaling_r > signaling_m:
                    dict_segment['Sinalização.R'] = ["X"]
                    fc_signaling = 1
                elif signaling_b == 0 and signaling_m == 0 and signaling_r == 0:
                    dict_segment['Sinalização.B'] = ["X"]
                    fc_signaling = 0.25

                dict_segment['ICC'] = [(30*fc_mowing) + (20*fc_drainage) + (50*fc_signaling)]
                dict_segment['ICM'] = dict_segment['ICP'][0]*0.7 + dict_segment['ICC'][0]*0.3
            
                # Concatenating final df
                df_final = pd.concat([df_final, pd.DataFrame.from_dict(dict_segment)], ignore_index=True)

        return df_final, file_header

    def Export(self, df_final, file_header):
        # Model Name
        modelo_path = r'C:\Pavesys\Templates\Padrão PAVESYS\ICM.xlsx'

        # Coping df
        planilha = df_final.copy(deep=True)

        # Abrindo arquivo modelo
        new_name = "ICM " + (os.path.basename(self.file_path).split(".")[0]).replace("PP_", "") + ".xlsx"
        print("Iniciando exportação do arquivo: " + new_name)
        nome_file = os.path.join(os.path.dirname(self.file_path), new_name)
        shutil.copy(modelo_path, nome_file)
        book = op.load_workbook(nome_file)

        # Editando arquivo
        sheet = book["_"]
        start_col = 2
        start_row = 11
        end_col = 32

        # Copia formatação
        for line in range(planilha.shape[0]):
            for col in range(end_col):
                sheet.cell(row = start_row + line, column = start_col + col)._style = copy(sheet.cell(row = start_row, column = col+2)._style)
                sheet.row_dimensions[start_row + line].height = 15

        # Filling in data
        start_col = 4
        anexo_colunas = ["Rodovia", "Início", "Fim", "Extensão", "P Alto", "P Médio", "P Baixo", "R Alto", "R Médio", "R Baixo", "Tr Alto", "Tr Médio", "Tr Baixo",
                         "Roçada.B", "Roçada.M", "Roçada.R", "Drenagem.B", "Drenagem.M", "Drenagem.R", "Sinalização.B", "Sinalização.M", "Sinalização.R", "Data",
                         "Latitude", "Longitude", "Observação", "ICC", "ICP", "ICM"]
        for line in range(planilha.shape[0]):
            for col in range(len(anexo_colunas)):
                sheet.cell(row = start_row + line, column = start_col + col).value = planilha[anexo_colunas[col]][line]
            
        # Filling in Header
        sheet.cell(row = 5, column = 2).value = "Rodovia: " + str(file_header["Road"])
        sheet.cell(row = 6, column = 2).value = "Trecho: " + str(file_header["STH"])
        sheet.cell(row = 7, column = 2).value = "Pista: " + str(file_header["Track"])
        sheet.cell(row = 5, column = 11).value = "Sentido: " + str(file_header["Way"])
        sheet.cell(row = 6, column = 11).value = "Faixa: " + str(file_header["Lane"])
        sheet.cell(row = 7, column = 11).value = "Data: " + str(file_header["Data"])
        sheet.cell(row = 5, column = 21).value = "Início (km): " + str(file_header["km ini"])
        sheet.cell(row = 6, column = 21).value = "Fim (km): " + str(file_header["km fim"])
        sheet.cell(row = 5, column = 30).value = str(file_header["one way"])
        sheet.cell(row = 6, column = 30).value = str(file_header["two way"])

        sheet.print_area = 'A1:AG' + str(planilha.shape[0] + start_row - 1)
        book.save(nome_file)
        print("    ... pronto para uso \n")

    def CallICM(self):
        self.list_df_main = self.SeparaDF(pd.read_excel(self.file_path, header=7))
        #print('ok')

        if os.path.basename(self.file_path).__contains__("_1."):
            df_final, file_header = self.Processing()
            #print(df_final)
        elif os.path.basename(self.file_path).__contains__("_RAMO"):
            list_ramo = []
            for df in self.list_df_main:
                df['Início'] = df['Início'] - df['Início'].min()
                df['Fim'] = df['Fim'] - df['Fim'].min()
                list_ramo.append(df)
            self.list_df_main = list_ramo
            df_final, file_header = self.Processing()
            #print(df_final)
        else:
            split_name = os.path.basename(self.file_path).split("_")
            base_name = split_name[0] + "_" + split_name[1] + "_" + split_name[2] + "_" + split_name[3] + "_1."
            for file in os.listdir(os.path.dirname(self.file_path)):
                if file.__contains__(base_name):
                    df_base = pd.read_excel(os.path.join(os.path.dirname(self.file_path), file), header=7)
                    df_final, file_header = self.Processing(df_base)
                    #print(df_final)
        #print(df_final)
        self.Export(df_final, file_header)

def main():
    #path = r'C:\Users\Pavesys - MAQ70\Desktop\python\ICM'
    path = filedialog.askdirectory()
    print(path)
    step = 1 #Segment in km

    excel_paths = []
    for video in os.listdir(path):
        if video.endswith('.xlsx') or video.endswith('.xls'):
            if video.startswith('PP_'):
                if not video.__contains__("_ATR.") and not video.__contains__("_IRI."):
                    if not video.startswith('ICM '):
                        excel_paths.append(os.path.join(path, video))

    for sheet_path in excel_paths:
        print(sheet_path)
        icm = CalcICM(sheet_path, step)
        icm.CallICM()
    messagebox.showinfo('Sucesso', 'Arquivos gerados com sucesso!')

if __name__ == "__main__":
    main()

